import tkinter as tk
from tkinter import messagebox
import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active  # 假设我们操作的是活动的工作表

# 定义增删函数

def add_record():
    # 获取用户输入
    title = title_entry.get()
    author = author_entry.get()
    # ...获取其他字段...

    # 添加到工作表
    ws.append([title, author, '', '', '', '', '', '', '', ''])
    wb.save('data.xlsx')  # 保存更改
    messagebox.showinfo("成功", "记录已添加")

def delete_record():
    # 通过标题删除记录
    title_to_delete = title_entry.get()
    rows_to_delete = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == title_to_delete:
                rows_to_delete.append(cell.row)

    for row in reversed(rows_to_delete):  # 反向删除，避免索引变化
        ws.delete_rows(row)

    wb.save('data.xlsx')
    messagebox.showinfo("成功", "记录已删除")

# 创建窗口
root = tk.Tk()
root.title("区块链文献管理")

# 创建输入框
title_label = tk.Label(root, text="标题")
title_label.pack()
title_entry = tk.Entry(root)
title_entry.pack()

author_label = tk.Label(root, text="作者")
author_label.pack()
author_entry = tk.Entry(root)
author_entry.pack()

# ...为其他字段创建输入框...

# 创建按钮
add_button = tk.Button(root, text="添加记录", command=add_record)
add_button.pack()

delete_button = tk.Button(root, text="删除记录", command=delete_record)
delete_button.pack()

def find_record():
    # 通过关键词查找记录
    keyword = title_entry.get()
    found_records = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=10):  # 假设我们有10列
        for cell in row:
            if keyword in str(cell.value):
                # 将找到的记录添加到列表中
                found_records.append(f"行 {cell.row}: {', '.join(str(col.value) for col in row)}")
                break  # 跳出循环，避免重复添加同一行的记录

    if found_records:
        messagebox.showinfo("找到记录", "\n".join(found_records))
    else:
        messagebox.showinfo("未找到", "没有找到包含该关键词的记录")

def update_record():
    # 通过标题更新记录
    title_to_update = title_entry.get()
    new_author = author_entry.get()
    # ...获取其他字段...

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == title_to_update:
                # 更新整行的数据
                ws.cell(row=cell.row, column=2, value=new_author)
                # ...更新其他字段...
                wb.save('data.xlsx')
                messagebox.showinfo("成功", "记录已更新")
                return
    messagebox.showinfo("未找到", "没有找到该标题的记录")

# ...添加查询和修改按钮...

find_button = tk.Button(root, text="查找记录", command=find_record)
find_button.pack()

update_button = tk.Button(root, text="更新记录", command=update_record)
update_button.pack()

# 运行主循环
root.mainloop()
